'''
项目介绍：爬取中国传媒大学新闻报告讲座，链接存在CUC.xlsx中，文本存在CUC_Text.zip中
说明：
1.起始网页：http://www.cuc.edu.cn/bgjz/list1.htm
'''
import requests
import lxml
import re
import time 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def writeText(title,text):
    '''
    将文本写入文件夹中
    title为文件名
    '''  
    fp=open(title+".txt","w",encoding='utf-8')
    fp.write(text)
    fp.close()

def getText(url):
    '''
    返回网页文案
    '''
    Text=''
    Text=Text+url+"\n"
    html=requests.get(url)
    html.encoding='utf-8'
    bs=BeautifulSoup(html.text,'lxml')
    for p in bs.find('article').findAll('p'):
        Text=Text+"\n"+p.get_text()
    return Text

def getURL(seedURL):
    '''
    返回从起始页开始的所有内链
    '''
    global postPath
    global char_set

    url_list=[]
    for i in range(1,39):
        url=seedURL.format(index=i)

        html=requests.get(url)
        html.encoding='utf-8'
        bs=BeautifulSoup(html.text,'lxml')

        for item in bs.findAll('h3',{'class':'tit'}):
            link=item.find('a')
            Title=link['title']
            for char in char_set:   #用空格替换保留字符
                Title=Title.replace(char,' ')
            Address=postPath+link['href']
            url_list.append([Title,Address])
    time.sleep(1)
    return url_list

postPath="http://www.cuc.edu.cn"  #前缀网址
char_set=['/','\\',':','*','?','|','"','>','<'] #保留字符
if __name__=="__main__":
    path = '***/Desktop'
    os.mkdir(path + './CUC_Text')
    path='***/CUC_Text'
    os.chdir(path)

    seedURL="http://www.cuc.edu.cn/bgjz/list{index}.htm"
    url_list=getURL(seedURL)

    try:
        for i in url_list:
            writeText(i[0],getText(i[1])) 
            time.sleep(1)
    finally:
        wb=Workbook()
        sheet=wb.active
        for link,index in zip(url_list,range(1,len(url_list)+1)):
            sheet.cell(index,1).value=link[0]
            sheet.cell(index,2).value=link[1] 
        wb.save("***/CUC.xlsx")
