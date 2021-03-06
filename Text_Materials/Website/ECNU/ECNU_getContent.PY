'''
项目介绍：爬取华东师范大学传播学院所有讲座预告，链接存在ECNU.xlsx中，文本存在ECNU_Text.zip中
说明：
1.起始网页：http://www.comm.ecnu.edu.cn/htmlaction.do?method=toGetSubNewsList&menuType=7&pageNo=0
2.访问内容时需要cookie
3.该网站源有点问题，大部分文章失效
'''
import requests
import lxml
import re
import time 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def writeText(title,text):
    '''
    将文本写入文件夹中
    title为文件名
    '''  
    fp=open(title+".txt","w",encoding='utf-8')
    fp.write(text)
    fp.close()

def getText(url):
    '''
    返回网页文案
    '''
    Text=''
    Text=Text+url+"\n"
    #设置请求头
    session=requests.Session()
    headers={
        'User-Agent':'Mozilla/5.0 (Macintosh;Intel Mac OS X 10_9_5)'
        'AppleWebKit 537.36 (KHTML,like Gecko) Chrome',
        'Accept':'text/html,application/xhtml+xml,application/xml;'
        'q=0.9,image/webp,*/*;q=0.8',
        'Cookie': '***' #浏览器抓包得
    }
    html=session.get(url,headers=headers)
    bs=BeautifulSoup(html.text,'lxml')
    content=bs.find('div',{'id':'news_detail_content'})
    if content==None:return None  #即错误页面
    return Text+content.get_text()

def getURL(seedURL):
    '''
    由起始页面获取所有内链
    '''
    global postPath
    global char_set

    url_list=[]
    for i in range(0,6):
        url=seedURL.format(index=i)
        html=requests.get(url)
        bs=BeautifulSoup(html.text,'lxml')
        content=bs.find('div',{'class':'news_content'})
        for link in content.findAll('a',href=re.compile('^(/htmlaction\.do\?method=toHtmlDetail.*)')):
            Title=link.get_text()[1:][:-1]   #去掉第一个和最后一个换行符
            for char in char_set:   #用空格替换保留字符
                Title=Title.replace(char,' ')
            Address=postPath+link['href']
            url_list.append([Title,Address])
        time.sleep(1)
    return url_list

postPath="http://www.comm.ecnu.edu.cn"  #前缀网址
char_set=['/','\\',':','*','?','|','"','>','<'] #保留字符
if __name__=="__main__":
    path = '***'
    os.mkdir(path + './ECNU_Text')
    path='***/ECNU_Text'
    os.chdir(path)

    seedURL="http://www.comm.ecnu.edu.cn/htmlaction.do?method=toGetSubNewsList&menuType=7&pageNo={index}"
    url_list=getURL(seedURL)

    try:
        for i in url_list:
            Text=getText(i[1])
            if(Text==None):i[0]="错误网页!"+i[0]  #在xlsx中标注错误页面
            else:writeText(i[0],Text)
        #第一次获取时该网站就封了cookie，对爬虫的识别应该比较敏感
            time.sleep(3)   
    finally:  #最后记录一下所有链接
        wb=Workbook()
        sheet=wb.active
        for link,index in zip(url_list,range(1,len(url_list)+1)):
            sheet.cell(index,1).value=link[0]
            sheet.cell(index,2).value=link[1]
        wb.save("***/ECNU.xlsx")
